from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
import os.path
import requests

url = "https://iotwebserver/api/get-device"

class room:
    roomName = ""
    roomTempt = ""
    temptTime = ""

def main():
    data = get_data(url)

    time = datetime.datetime.now()
    formattedTime = time.strftime("%Y-%m")
    date = time.strftime("%d")
    print(formattedTime)

    for elm in data:
        excelName = elm.roomName + " " + formattedTime + ".xlsx"
        # If the Excel file with the approriate date excists, we just update the excel file. Else, we create a new one with the approriate date name.
        if os.path.isfile("./"+excelName):
            print("It's real!")
            update_excel(elm.roomTempt,elm.temptTime,excelName,date)
        else:
            print("nope")
            create_excel(elm.roomTempt,elm.temptTime,excelName,date)
    #return os.path.abspath(excelName)

# Gets all the data we need from the entire list of devices from the URL.
def get_data(url) -> list:
    roomList = []
    # Where our values from the get request will be stored
    # We set "verify" to be false, so that we dont have to worry about the certificate.
    r = requests.get(url, verify=False)
    # print(f"Response: {r.json()}")
    result = r.json()

    # If the name has the work "Temperature" in it, we create a new room variable and stores the data
    for entry in result:
        print(result.get(entry).get("name"))
        if result.get(entry).get("name").endswith("Temperature"):
            temptRoom = room() #Remember the "()", to make sure that we initialize a new one each time
            temptRoom.roomName = result.get(entry).get("name")
            temptRoom.roomTempt = result.get(entry).get("data").get("temperature").get("value")
            temptRoom.temptTime = result.get(entry).get("data").get("temperature").get("time")
            roomList.append(temptRoom)
            
    for i in roomList:
        print(i.roomName)
        print(i.roomTempt)
        print(i.temptTime)
    
    return roomList

# Creats a new excel file
def create_excel(tempt, time,exName,date):
    # Creates the new excel file
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active
    # Changes the name from the default "Sheet" to the approoriate date
    ws.title = date

    # Formats time to be "HH:MM" hours and minutes
    date_time = datetime.datetime.fromtimestamp(time)
    formattedTime = date_time.strftime('%H:%M')

    # We convert tempt to float, otherwise the charts in the excel files will be fucked.
    tempt = float(tempt)
    # Data can be assigned directly to cells
    ws.append([formattedTime,tempt])

    # Save the file
    wb.save(exName)
    return

# Updates an excisting excel file
def update_excel(tempt,time,exName,date):
    # Checks if the sheet with the approriate date excists. If it does not, we create the new sheet
    try:
        wb = load_workbook(exName)
        ws = wb[date]
    except Exception as e:
        print(e)
        print("Creating new sheet")
        create_excel_sheet(exName,date)
    
    # Loads the approriate excel file and worksheet
    wb = load_workbook(exName)
    ws = wb[date]

    # Formats time to be equal to 
    date_time = datetime.datetime.fromtimestamp(time)
    formattedTime = date_time.strftime('%H:%M')

    # We convert tempt to float, otherwise the charts in the excel files will be fucked.
    tempt = float(tempt)
    # Data can be assigned directly to cells
    ws.append([formattedTime,tempt])

    # Save the file
    wb.save(exName)
    return

# Creates a new sheet with the approriate date, incase the excel file does not already have that
def create_excel_sheet(exName,date):
    wb = load_workbook(exName)
    wb.create_sheet(date)
    wb.save(exName)
    return
    
# Excists so that we run the program normally
if __name__=="__main__":
    main()